from openpyxl import load_workbook, Workbook

caminho_planilha = "planilhas/data.xlsx"

planilha = load_workbook(caminho_planilha)

dados_servidores = planilha.active

# Lista com o nome dos servidores que deseja em encontrar
lista_servidores = []

# Verificando lotação e exercicio dos servidores que estão em lista_servidores
dados_retorno = []
for servidores in dados_servidores.iter_rows(min_row=2, min_col=1, max_col=4):
    nome = servidores[0].value
    lotacao = servidores[2].value
    exercicio = servidores[3].value

    if nome in lista_servidores:
        dados_retorno.append({"Nome": nome, "Lotação": lotacao, "Exercício": exercicio})

# ? Obtendo nome dos servidores que nao foram encontrados na planilha
nao_encontrado = []
for nome in lista_servidores:
    # Verifica se o nome está em dados_retorno
    if not any(dado["Nome"] == nome for dado in dados_retorno):
        nao_encontrado.append(nome)
# * Exibindo os nomes não encontrados
print("Servidores não encontrados:", nao_encontrado)


# ! Criando nova Planilha com os dados retornados dos servidores
nova_planilha = Workbook()
nova_aba = nova_planilha.active
nova_aba.title = "Dados Servidores"

# Criando Cabeçalho
nova_aba.append(["Nome", "Lotação", "Exercício"])

# * Adicionando os dados na Planilha
for dado in dados_retorno:
    nova_aba.append([dado["Nome"], dado["Lotação"], dado["Exercício"]])

# ! Salvando a nova planilha
caminho_nova_planilha = "planilhas/dados_servidores.xlsx"
nova_planilha.save(caminho_nova_planilha)

print(f"Planilha criada com sucesso em {caminho_nova_planilha}")
